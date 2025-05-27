import os
from sentence_transformers import SentenceTransformer
from autoquest.config import config
import openpyxl
from chromadb import PersistentClient
import json

def embed_documents(data_dir="data"):
    """
    Read all .txt and .xlsx files in the data directory, generate embeddings,
    and store them in ChromaDB.
    """
    # Load embedding model
    embedder = SentenceTransformer(config["embedding_model"])

    # Init ChromaDB
    client = PersistentClient(path=config["chroma_db_dir"])

    try:
        client.delete_collection("autoquest_knowledge")
    except:
        pass

    collection = client.create_collection(name="autoquest_knowledge")

    # Remove old collection if it exists
    try:
        client.delete_collection(collection)
    except Exception:
        pass

    collection = client.get_or_create_collection(name=config["chroma_collection"])

    # Loop through files
    for filename in os.listdir(data_dir):
        path = os.path.join(data_dir, filename)

        if filename.endswith(".txt"):
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
                emb = embedder.encode(content).tolist()
                collection.add(documents=[content], embeddings=[emb], ids=[filename])
                print(f"✅ Embedded TXT: {filename}")

        elif filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(path)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                question, answer, *_ = row
                if question and answer:
                    qa_text = f"Question: {question.strip()}\nAnswer: {answer.strip()}"
                    emb = embedder.encode(qa_text).tolist()
                    row_id = f"{filename}_{question[:30].strip()}"
                    collection.add(documents=[qa_text], embeddings=[emb], ids=[row_id])
            print(f"✅ Embedded XLSX: {filename}")

    print("✅ All documents embedded and saved to ChromaDB.")

def embed_qa_sheet_incrementally(sheet_path="data/QA_sheet.xlsx"):
    from chromadb import PersistentClient

    print("🔁 Checking for new Q&A entries to embed...")
    embedder = SentenceTransformer(config["embedding_model"])
    client = PersistentClient(path=config["chroma_db_dir"])
    collection = client.get_or_create_collection(name=config["chroma_collection"])

    log_path = os.path.join(config["chroma_db_dir"], "embedded_ids.json")
    if os.path.exists(log_path):
        with open(log_path, "r") as f:
            embedded_ids = set(json.load(f))
    else:
        embedded_ids = set()

    if not os.path.exists(sheet_path):
        print("⚠️ QA_sheet.xlsx not found. Skipping embedding.")
        return

    wb = openpyxl.load_workbook(sheet_path)
    sheet = wb.active
    new_count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        question, answer, *_ = row
        if question and answer:
            qa_id = f"{question.strip().lower()[:100]}"  # Safe ID
            if qa_id not in embedded_ids:
                qa_text = f"[Q] {str(question).strip()}\n[A] {str(answer).strip()}"
                emb = embedder.encode(qa_text).tolist()
                collection.add(documents=[qa_text], embeddings=[emb], ids=[qa_id])
                embedded_ids.add(qa_id)
                new_count += 1

    with open(log_path, "w") as f:
        json.dump(list(embedded_ids), f)

    print(f"✅ {new_count} new Q&A entries embedded.")

